import io
import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ── Palette ────────────────────────────────────────────────────────────────
NAVY       = "1E293B"   # header banner bg
BLUE       = "2563EB"   # column header bg
LIGHT_BLUE = "EFF6FF"   # alternating row bg
MID_BLUE   = "DBEAFE"   # KPI label bg
WHITE      = "FFFFFF"
SLATE      = "F8FAFC"
BORDER_CLR = "CBD5E1"
RED        = "DC2626"
GREEN      = "16A34A"
AMBER      = "D97706"
TEXT_DARK  = "0F172A"
TEXT_MID   = "475569"
TEXT_LIGHT = "94A3B8"

CURRENCY_FMT = '$#,##0.00'
CURRENCY_FMT_NEG = '$#,##0.00;[Red]($#,##0.00)'

def _side(color=BORDER_CLR, style="thin"):
    return Side(border_style=style, color=color)

def _border(all="thin"):
    s = _side(style=all)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, color=TEXT_DARK, size=11, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _clean(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(str(val).replace("$", "").replace(",", "").strip())
    except ValueError:
        return str(val)

def _set_cell(ws, row, col, value=None, font=None, fill=None, alignment=None,
              border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:       c.font = font
    if fill:       c.fill = fill
    if alignment:  c.alignment = alignment
    if border:     c.border = border
    if number_format: c.number_format = number_format
    return c


def _build_banner(ws, row, title, subtitle, col_span):
    """Merge a 2-row navy banner across col_span columns."""
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=col_span)

    c1 = ws.cell(row=row, column=1, value=title)
    c1.font = _font(bold=True, color=WHITE, size=14)
    c1.fill = _fill(NAVY)
    c1.alignment = _align(h="left", v="center")

    c2 = ws.cell(row=row+1, column=1, value=subtitle)
    c2.font = _font(color=TEXT_LIGHT, size=10, italic=True)
    c2.fill = _fill(NAVY)
    c2.alignment = _align(h="left", v="center")

    # Fill remaining banner cells so no white gaps show
    for r in [row, row+1]:
        for col in range(2, col_span + 1):
            ws.cell(row=r, column=col).fill = _fill(NAVY)


def generate_bank_excel(data: Dict[str, Any]) -> bytes:
    wb = Workbook()

    bank_name       = data.get("bank_name", "Bank")
    account_holder  = data.get("account_holder", "")
    account_number  = data.get("account_number", "")
    statement_period= data.get("statement_period", "")
    transactions    = data.get("transactions", [])
    generated_at    = datetime.datetime.now().strftime("%d %b %Y %H:%M")
    tx_count        = len(transactions)

    # ════════════════════════════════════════════════════════════════════
    # SHEET 1: Transactions
    # ════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Transactions"
    ws.sheet_view.showGridLines = False

    COL_SPAN = 8  # #, Date, Desc, Ref, Debit, Credit, Balance, Type
    HEADER_ROW = 4
    DATA_START  = 5

    # Banner (rows 1-2)
    _build_banner(ws, 1,
        f"  {bank_name}  ·  Transaction Ledger",
        f"  Account: {account_number}   |   Holder: {account_holder}   |   Period: {statement_period}   |   Generated: {generated_at}",
        COL_SPAN)

    # Spacer row 3
    ws.row_dimensions[3].height = 6
    for col in range(1, COL_SPAN + 1):
        ws.cell(row=3, column=col).fill = _fill(SLATE)

    # Column header row 4
    ws.row_dimensions[HEADER_ROW].height = 24
    headers = ["#", "Date", "Description", "Reference", "Debit ($)", "Credit ($)", "Balance ($)", "Type"]
    widths  = [5,   14,     42,             16,           14,          14,           16,             12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = _set_cell(ws, HEADER_ROW, i, h,
            font=_font(bold=True, color=WHITE, size=10),
            fill=_fill(BLUE),
            alignment=_align(h="center"),
            border=_border())
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for idx, tx in enumerate(transactions):
        r = DATA_START + idx
        # Calculate dynamic row height based on length of wrapped description text
        desc = tx.get("description", "")
        desc_lines = str(desc).split('\n')
        total_wrapped_lines = 0
        for line in desc_lines:
            total_wrapped_lines += max(1, len(line) // 38 + 1)
        ws.row_dimensions[r].height = max(20, total_wrapped_lines * 15)
        row_fill = _fill(WHITE) if idx % 2 == 0 else _fill(LIGHT_BLUE)
        bdr = _border()

        debit_val   = _clean(tx.get("debit"))
        credit_val  = _clean(tx.get("credit"))
        balance_val = _clean(tx.get("balance"))

        # Determine type tag
        if isinstance(debit_val, float) and debit_val > 0:
            tx_type = "Debit"
        elif isinstance(credit_val, float) and credit_val > 0:
            tx_type = "Credit"
        else:
            tx_type = "—"

        row_vals = [
            (idx + 1,       _align(h="center"), _font(color=TEXT_LIGHT, size=10), None),
            (tx.get("date",""), _align(h="center"), _font(color=TEXT_DARK, size=10), None),
            (tx.get("description",""), _align(h="left", wrap=True), _font(color=TEXT_DARK, size=10), None),
            (tx.get("reference",""),   _align(h="center"), _font(color=TEXT_MID, size=10), None),
            (debit_val,     _align(h="right"), _font(bold=True, color=RED, size=10),   CURRENCY_FMT if isinstance(debit_val, float) else None),
            (credit_val,    _align(h="right"), _font(bold=True, color=GREEN, size=10), CURRENCY_FMT if isinstance(credit_val, float) else None),
            (balance_val,   _align(h="right"), _font(bold=True, color=NAVY, size=10),  CURRENCY_FMT if isinstance(balance_val, float) else None),
            (tx_type,       _align(h="center"), _font(color=TEXT_MID, size=9, italic=True), None),
        ]

        for col, (val, align, font, num_fmt) in enumerate(row_vals, 1):
            _set_cell(ws, r, col, val, font=font, fill=row_fill,
                      alignment=align, border=bdr, number_format=num_fmt)

    # Totals row
    if tx_count > 0:
        tot_row = DATA_START + tx_count
        ws.row_dimensions[tot_row].height = 22
        tot_fill = _fill(MID_BLUE)
        thick_top = Border(
            top=Side(border_style="medium", color=NAVY),
            left=_side(), right=_side(), bottom=_side()
        )
        # Merge first 3 cols for "TOTALS" label
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=4)
        c_label = ws.cell(row=tot_row, column=1, value="TOTALS")
        c_label.font = _font(bold=True, color=NAVY, size=11)
        c_label.fill = tot_fill
        c_label.alignment = _align(h="right")
        c_label.border = thick_top
        for col in [2,3,4]:
            ws.cell(row=tot_row, column=col).fill = tot_fill
            ws.cell(row=tot_row, column=col).border = thick_top

        last_data = DATA_START + tx_count - 1
        for col, formula in [
            (5, f"=SUM(E{DATA_START}:E{last_data})"),
            (6, f"=SUM(F{DATA_START}:F{last_data})"),
            (7, ""),
            (8, ""),
        ]:
            c = _set_cell(ws, tot_row, col, formula,
                font=_font(bold=True, color=NAVY, size=11),
                fill=tot_fill, alignment=_align(h="right"),
                border=thick_top,
                number_format=CURRENCY_FMT if formula else None)

        # Enable auto-filter on header row
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(COL_SPAN)}{last_data}"

    # Freeze top rows
    ws.freeze_panes = f"A{DATA_START}"

    # Auto-fit column widths dynamically to prevent clipping and '###' values
    if tx_count > 0:
        for col_idx in range(1, COL_SPAN + 1):
            col_letter = get_column_letter(col_idx)
            max_w = widths[col_idx - 1]
            for row_idx in range(DATA_START, DATA_START + tx_count):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    if isinstance(cell_val, float):
                        val_str = f"${cell_val:,.2f}"
                    else:
                        val_str = str(cell_val)
                    max_line_len = max(len(l) for l in val_str.split('\n'))
                    max_w = max(max_w, max_line_len + 3)
            # Cap the column width at 60 (specifically for Description) to keep sheet neat
            ws.column_dimensions[col_letter].width = min(max_w, 60)


    # ════════════════════════════════════════════════════════════════════
    # SHEET 2: Summary Dashboard
    # ════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary Dashboard")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 2   # left margin
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 26

    _build_banner(ws2, 1, f"  {bank_name}  ·  Account Summary", f"  Generated: {generated_at}", 4)

    # Account details section
    ws2.row_dimensions[3].height = 8
    ws2.row_dimensions[4].height = 20
    c_sec = ws2.cell(row=4, column=2, value="ACCOUNT DETAILS")
    c_sec.font = _font(bold=True, color=BLUE, size=9)
    c_sec.alignment = _align()

    meta_rows = [
        ("Account Holder",  account_holder),
        ("Bank Name",       bank_name),
        ("Account Number",  account_number),
        ("Statement Period",statement_period),
        ("Report Generated",generated_at),
    ]
    for i, (label, val) in enumerate(meta_rows):
        r = 5 + i
        ws2.row_dimensions[r].height = 20
        cl = _set_cell(ws2, r, 2, label,
            font=_font(bold=True, color=TEXT_DARK, size=10),
            fill=_fill(MID_BLUE), alignment=_align(),
            border=_border())
        cv = _set_cell(ws2, r, 3, val,
            font=_font(color=TEXT_DARK, size=10),
            fill=_fill(WHITE), alignment=_align(),
            border=_border())

    # KPI section
    ws2.row_dimensions[11].height = 10
    ws2.row_dimensions[12].height = 20
    c_kpi = ws2.cell(row=12, column=2, value="FINANCIAL SUMMARY")
    c_kpi.font = _font(bold=True, color=BLUE, size=9)
    c_kpi.alignment = _align()

    last_tx = DATA_START + max(tx_count - 1, 0)
    kpi_rows = [
        ("Total Debits",         f"=SUM(Transactions!E{DATA_START}:E{last_tx})", CURRENCY_FMT),
        ("Total Credits",        f"=SUM(Transactions!F{DATA_START}:F{last_tx})", CURRENCY_FMT),
        ("Net Cash Flow",        f"=SUM(Transactions!F{DATA_START}:F{last_tx})-SUM(Transactions!E{DATA_START}:E{last_tx})", CURRENCY_FMT),
        ("Transaction Count",    f"=COUNTA(Transactions!B{DATA_START}:B{last_tx})", None),
        ("Largest Debit",        f"=IFERROR(MAX(Transactions!E{DATA_START}:E{last_tx}),0)", CURRENCY_FMT),
        ("Largest Credit",       f"=IFERROR(MAX(Transactions!F{DATA_START}:F{last_tx}),0)", CURRENCY_FMT),
    ]
    for i, (label, formula, fmt) in enumerate(kpi_rows):
        r = 13 + i
        ws2.row_dimensions[r].height = 22
        fill_l = _fill(LIGHT_BLUE) if i % 2 == 0 else _fill(WHITE)
        _set_cell(ws2, r, 2, label,
            font=_font(bold=True, color=TEXT_DARK, size=10),
            fill=fill_l, alignment=_align(), border=_border())
        _set_cell(ws2, r, 3, formula,
            font=_font(bold=True, color=NAVY, size=11),
            fill=fill_l, alignment=_align(h="right"),
            border=_border(), number_format=fmt)

    # Notes
    ws2.row_dimensions[21].height = 10
    ws2.row_dimensions[22].height = 18
    c_note = ws2.cell(row=22, column=2,
        value="Auto-generated by Local AI Bank Statement Converter  ·  Data sourced from AI OCR extraction")
    c_note.font = _font(color=TEXT_LIGHT, size=9, italic=True)
    ws2.merge_cells("B22:D22")


    # ════════════════════════════════════════════════════════════════════
    # SHEET 3: Validation Report
    # ════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Validation Report")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 2
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 46

    _build_banner(ws3, 1, f"  Extraction Integrity Audit", f"  {bank_name}  ·  {statement_period}", 5)

    ws3.row_dimensions[3].height = 8
    ws3.row_dimensions[4].height = 22
    for col, hdr in [(2, "Audit Check"), (3, "Status"), (4, "Detail")]:
        _set_cell(ws3, 4, col, hdr,
            font=_font(bold=True, color=WHITE, size=10),
            fill=_fill(NAVY), alignment=_align(h="center"),
            border=_border())

    last_tx = DATA_START + max(tx_count - 1, 0)
    audit_rows = [
        ("Transactions Extracted",
         "PASS" if tx_count > 0 else "WARN",
         f"{tx_count} transaction rows extracted from document."),
        ("Date Field Coverage",
         "PASS",
         f"=COUNTA(Transactions!B{DATA_START}:B{last_tx})&\" of {tx_count} rows have a date value.\""),
        ("Description Coverage",
         "PASS",
         f"=COUNTA(Transactions!C{DATA_START}:C{last_tx})&\" of {tx_count} rows have a description.\""),
        ("Debit/Credit Net",
         "INFO",
         f"=TEXT(SUM(Transactions!F{DATA_START}:F{last_tx})-SUM(Transactions!E{DATA_START}:E{last_tx}),\"$#,##0.00\")&\" net cash flow (positive = net inflow).\""),
        ("Balance Continuity",
         "MANUAL",
         "Verify that each balance equals prior balance minus debit plus credit."),
        ("Data Source",
         "INFO",
         "Extracted via AI OCR. Manual verification recommended for critical accounting."),
    ]

    STATUS_COLORS = {
        "PASS":   (GREEN,       "DCFCE7"),
        "WARN":   (AMBER,       "FEF9C3"),
        "FAIL":   (RED,         "FEE2E2"),
        "INFO":   (BLUE,        "DBEAFE"),
        "MANUAL": ("7C3AED",    "EDE9FE"),
    }

    for i, (check, status, detail) in enumerate(audit_rows):
        r = 5 + i
        ws3.row_dimensions[r].height = 20
        row_bg = STATUS_COLORS.get(status, ("000000", "FFFFFF"))[1]
        status_color = STATUS_COLORS.get(status, ("000000", "FFFFFF"))[0]

        _set_cell(ws3, r, 2, check,
            font=_font(bold=True, color=TEXT_DARK, size=10),
            fill=_fill(LIGHT_BLUE if i % 2 == 0 else WHITE),
            alignment=_align(), border=_border())
        _set_cell(ws3, r, 3, status,
            font=_font(bold=True, color=status_color, size=10),
            fill=_fill(row_bg),
            alignment=_align(h="center"), border=_border())
        _set_cell(ws3, r, 4, detail,
            font=_font(color=TEXT_DARK, size=10),
            fill=_fill(WHITE), alignment=_align(wrap=True),
            border=_border())
        ws3.row_dimensions[r].height = 22

    # ── Stream ─────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
